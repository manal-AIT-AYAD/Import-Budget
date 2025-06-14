from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from datetime import datetime
import re

def process_budget_excel(source_path: str, output_path: str = "compte_de_resultats_budget1.xlsx") -> str:
    workbook = load_workbook(source_path)
    sheet = workbook.active

    header_row_index = None
    for row in range(1, 20):
        code_cell = sheet.cell(row=row, column=1).value
        name_cell = sheet.cell(row=row, column=2).value

        if code_cell and name_cell:
            if str(code_cell).strip().lower() == "code" and (
                str(name_cell).strip().lower().startswith("nom") or
                "compte" in str(name_cell).strip().lower()
            ):
                header_row_index = row
                break

    if header_row_index is None:
        raise ValueError("Ligne d'en-tête non trouvée")

    solde_year = None
    solde_header_cell = sheet.cell(row=header_row_index, column=3).value
    if solde_header_cell:
        year_match = re.search(r'20\d{2}', str(solde_header_cell))
        if year_match:
            solde_year = int(year_match.group(0))

    if solde_year is None:
        for row in range(max(1, header_row_index - 5), header_row_index + 2):
            for col in range(1, 6):
                cell_value = str(sheet.cell(row=row, column=col).value or "")
                year_match = re.search(r'20\d{2}', cell_value)
                if year_match:
                    solde_year = int(year_match.group(0))
                    break
            if solde_year:
                break

    if solde_year is None:
        solde_year = datetime.now().year

    budget_year = solde_year + 1
    rows_to_delete = []
    for row in range(header_row_index + 1, sheet.max_row + 1):
        if not sheet.cell(row=row, column=1).value:
            rows_to_delete.append(row)
    for row in reversed(rows_to_delete):
        sheet.delete_rows(row)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True)
    centered_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")

    if "C1" in sheet.merged_cells:
        sheet.unmerge_cells("C1")
    sheet["C1"].value = None

    solde_col_letter = get_column_letter(3)
    solde_cell = sheet[f"{solde_col_letter}{header_row_index}"]
    solde_cell.value = f"Solde {solde_year}"
    solde_cell.font = Font(bold=True)
    solde_cell.alignment = centered_alignment

    year_cell = sheet.cell(row=header_row_index - 1, column=3)
    year_cell.value = budget_year
    year_cell.font = Font(bold=True, underline="single")
    year_cell.alignment = right_alignment

    last_data_row = header_row_index
    for row in range(header_row_index + 1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value:
            last_data_row = row

    new_columns = ["%"] + [
        "janvier", "février", "mars", "avril", "mai", "juin",
        "juillet", "août", "septembre", "octobre", "novembre", "décembre"
    ]
    start_col = 4
    total_col = start_col + len(new_columns)

    for i, title in enumerate(new_columns):
        col_letter = get_column_letter(start_col + i)
        cell = sheet[f"{col_letter}{header_row_index}"]
        cell.value = title
        cell.alignment = centered_alignment
        cell.border = thin_border
        cell.font = header_font

    total_col_letter = get_column_letter(total_col)
    cell = sheet[f"{total_col_letter}{header_row_index}"]
    cell.value = "Total"
    cell.alignment = centered_alignment
    cell.border = thin_border
    cell.font = header_font

    # Titre "Budget N+1" au-dessus des mois
    title_row_index = header_row_index - 1
    first_merge_col = get_column_letter(start_col)
    last_merge_col = get_column_letter(start_col + 12)
    sheet.merge_cells(f"{first_merge_col}{title_row_index}:{last_merge_col}{title_row_index}")
    cell = sheet[f"{first_merge_col}{title_row_index}"]

    # Format date et heure actuels
    now = datetime.now()
    formatted_date = now.strftime("%Y/%m/%d") 
    formatted_time = now.strftime("%H:%M") 

    # Appliquer le nouveau titre
    cell.value = f"Budget {budget_year} - {formatted_date} - {formatted_time}"
    cell.alignment = centered_alignment
    cell.fill = header_fill
    cell.font = Font(bold=True, size=14)
    cell.border = thin_border


    for col in range(1, start_col):
        cell = sheet.cell(row=header_row_index, column=col)
        cell.border = thin_border
        cell.font = header_font
        cell.alignment = centered_alignment

    for row in range(header_row_index + 1, last_data_row + 1):
        solde_cell_ref = f"${get_column_letter(3)}{row}"
        percent_cell_ref = f"${get_column_letter(start_col)}{row}"

        percent_cell = sheet.cell(row=row, column=start_col)
        percent_cell.value = 0
        percent_cell.number_format = '0.00%'
        percent_cell.alignment = centered_alignment
        percent_cell.border = thin_border

        for i in range(1, 13):
            monthly_col = start_col + i
            cell = sheet.cell(row=row, column=monthly_col)
            cell.value = f"=({solde_cell_ref} * {percent_cell_ref}) / 12"
            cell.number_format = '#,##0.00'
            cell.alignment = right_alignment
            cell.border = thin_border

        total_cell = sheet.cell(row=row, column=total_col)
        first_month_letter = get_column_letter(start_col + 1)
        last_month_letter = get_column_letter(start_col + 12)
        total_cell.value = f"=SUM({first_month_letter}{row}:{last_month_letter}{row})"
        total_cell.number_format = '#,##0.00'
        total_cell.alignment = right_alignment
        total_cell.border = thin_border

    for row in range(header_row_index + 1, last_data_row + 1):
        for col in range(1, start_col):
            cell = sheet.cell(row=row, column=col)
            cell.border = thin_border
            if col == 3:
                cell.number_format = '#,##0.00'
                cell.alignment = right_alignment
                sheet.column_dimensions[get_column_letter(col)].width = 15

    for col in range(1, total_col + 1):
        max_length = 0
        column = get_column_letter(col)
        for row in range(header_row_index, last_data_row + 1):
            cell_value = str(sheet.cell(row=row, column=col).value or "")
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        adjusted_width = max(max_length + 2, 12)
        if col == 3 or col >= start_col + 1:
            adjusted_width = max(adjusted_width, 15)
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save(output_path)
    return output_path
