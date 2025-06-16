import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
import unidecode

def transform_budget_data_append_sheet(input_files, existing_file, new_sheet_name="Import Odoo"):
    mois_liste = [
        'janvier', 'fevrier', 'mars', 'avril', 'mai', 'juin',
        'juillet', 'aout', 'septembre', 'octobre', 'novembre', 'decembre'
    ]
    mois_to_num = {mois: str(i+1).zfill(2) for i, mois in enumerate(mois_liste)}

    all_data_by_year = {}

    for input_file in input_files:
        print(f"Lecture du fichier source: {input_file}")
        try:
            df_raw = pd.read_excel(input_file, header=None)
        except Exception as e:
            print(f"Erreur lecture Excel: {e}")
            continue

        annee_budget = None
        for i in range(min(10, len(df_raw))):
            row = df_raw.iloc[i].astype(str)
            for val in row:
                match = re.search(r'(\d{4})', val)
                if match:
                    annee_candidate = int(match.group(1))
                    if 2000 <= annee_candidate <= 2100:
                        annee_budget = annee_candidate
                        break
            if annee_budget:
                break
        if not annee_budget:
            match = re.search(r'(\d{4})', input_file)
            annee_budget = int(match.group(1)) if match and 2000 <= int(match.group(1)) <= 2100 else datetime.now().year

        print(f"✅ Année détectée pour le fichier : {annee_budget}")

        header_row_idx = None
        for i in range(len(df_raw)):
            row_values = df_raw.iloc[i].astype(str).str.lower().str.strip()
            row_values = row_values.map(lambda x: unidecode.unidecode(x))
            if all(mois in row_values.values for mois in ['janvier', 'fevrier', 'mars']):
                header_row_idx = i
                break

        if header_row_idx is None:
            print(f"Aucune ligne d'en-tête détectée dans {input_file}")
            continue

        df_source = pd.read_excel(input_file, header=header_row_idx)

        if annee_budget not in all_data_by_year:
            all_data_by_year[annee_budget] = []

        if 'Code' not in df_source.columns:
            df_source.rename(columns={df_source.columns[0]: 'Code'}, inplace=True)
        if 'Nom du compte' not in df_source.columns:
            df_source.rename(columns={df_source.columns[1]: 'Nom du compte'}, inplace=True)

        colonnes_presentes = df_source.columns.map(lambda x: unidecode.unidecode(str(x)).lower().replace(" ", ""))
        mois_map = {}

        for mois in mois_liste:
            for col, col_clean in zip(df_source.columns, colonnes_presentes):
                if col_clean == mois:
                    mois_map[mois] = col
                    break

        print(f"Colonnes mois détectées pour {input_file} : {mois_map}")

        compteur_ligne = 1
        for _, row in df_source.iterrows():
            code = row.get('Code')
            if pd.isna(code):
                continue

            try:
                code_int = int(float(code))
            except:
                continue

            for mois in mois_liste:
                if mois not in mois_map:
                    continue

                col_mois = mois_map[mois]
                montant = row.get(col_mois)

                if pd.isna(montant):
                    continue

                if isinstance(montant, str):
                    montant = montant.replace('.', '').replace(',', '.').replace(' ', '')
                try:
                    montant_float = float(montant)
                except:
                    continue

                mois_num = mois_to_num[mois]
                date_budget = f"01/{mois_num}/{annee_budget}"

                new_row = {
                    'annee': annee_budget,
                    'compteur_ligne': compteur_ligne,
                    'code_compte': str(code_int),
                    'montant': -montant_float,
                    'mois': mois,
                    'date': date_budget
                }
                all_data_by_year[annee_budget].append(new_row)
                compteur_ligne += 1

    try:
        wb = load_workbook(existing_file)
    except Exception as e:
        print(f"Erreur chargement fichier existant: {e}")
        return

    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]
    ws = wb.create_sheet(title=new_sheet_name)

    headers = ['', 'name', 'id', 'item_ids/id', 'item_ids/date', 'item_ids/account', 'item_ids/amount', '']
    ws.append(headers)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, cell in enumerate(ws[1], 1):
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    column_widths = {'A': 10, 'B': 20, 'C': 20, 'D': 25, 'E': 12, 'F': 15, 'G': 15, 'H': 10}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    row_num = 2
    for annee_budget in sorted(all_data_by_year.keys()):
        year_data = all_data_by_year[annee_budget]
        if not year_data:
            continue

        #budget_name = f"Budget {annee_budget}"
        now = datetime.now()
        formatted_date = now.strftime("%Y/%m/%d") 
        formatted_time = now.strftime("%H:%M")
        budget_name = f"Budget {annee_budget} - {formatted_date} - {formatted_time}"

        budget_id = f"budget_{annee_budget}_00001"

        compteur_global = 1

        for mois in mois_liste:
            for item in year_data:
                if item['mois'] != mois:
                    continue

                if compteur_global == 1:
                    row_values = [
                        annee_budget,
                        budget_name,
                        budget_id,
                        f"lignes_budget_{annee_budget}{compteur_global}",
                        item['date'],
                        item['code_compte'],
                        item['montant'],
                        item['mois']
                    ]
                else:
                    row_values = [
                        "",
                        "",
                        "",
                        f"lignes_budget_{annee_budget}{compteur_global}",
                        item['date'],
                        item['code_compte'],
                        item['montant'],
                        item['mois']
                    ]

                ws.append(row_values)

                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.border = thin_border

                    if col_idx == 7:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
                    elif col_idx in [1, 6]:
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        cell.alignment = Alignment(horizontal='left')

                row_num += 1
                compteur_global += 1

    try:
        wb.save(existing_file)
        print(f"✅ Feuille '{new_sheet_name}' ajoutée dans : {existing_file}")
    except Exception as e:
        print(f"Erreur sauvegarde: {e}")


if __name__ == "__main__":
    input_files = ["compte_de_resultats_budget1.xlsx"]  
    output_file = "compte_de_resultats_budget1.xlsx"
    transform_budget_data_append_sheet(input_files, output_file)
